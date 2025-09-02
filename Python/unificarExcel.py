import openpyxl
from tkinter import filedialog, Tk

# Solicitar al usuario que seleccione un archivo Excel
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

# Cargar el archivo Excel
workbook = openpyxl.load_workbook(file_path)

# Crear una nueva hoja llamada "punion"
punion_sheet = workbook.create_sheet("punion")

# Buscar todas las tablas en el archivo
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    if len(sheet[1]) == 9:  # Validar que tenga 9 columnas
        # Copiar los registros a la tabla "union"
        for row in sheet.iter_rows(min_row=2, values_only=True):
            punion_sheet.append(row)

# Guardar los cambios en el archivo
workbook.save(file_path)

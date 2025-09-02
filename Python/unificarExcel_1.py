import openpyxl
from tkinter import filedialog, Tk, messagebox
import pandas as pd

# Función para convertir las fechas en formato string a objetos datetime
def parse_fecha(fecha_str):
    try:
        fecha_datetime = pd.to_datetime(fecha_str, format='%d/%m/%Y')
        if pd.notna(fecha_datetime):  # Verificar si la fecha es válida
            return fecha_datetime.strftime('%d/%m/%Y')
        else:
            return fecha_str  # Dejar la fecha original sin cambios
    except ValueError:
        return fecha_str  # Dejar la fecha original sin cambios

# Solicitar al usuario que seleccione un archivo Excel
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

# Cargar el archivo Excel
workbook = openpyxl.load_workbook(file_path)

# Eliminar la hoja "punion" si existe
if "punion" in workbook.sheetnames:
    workbook.remove(workbook["punion"])

# Crear una nueva hoja llamada "punion"
punion_sheet = workbook.create_sheet("punion")

# Agregar los encabezados a la tabla "union"
encabezados = ["tipo", "fecha", "descripcion", "valor", "capital", "cuotas", "pendiente", "tasaMV", "tasaEA"]
punion_sheet.append(encabezados)

# Buscar todas las tablas en el archivo
registros_unificados = set()  # Conjunto para evitar duplicados
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    if len(sheet[1]) == 9:  # Validar que tenga 9 columnas
        # Copiar los registros a la tabla "union" sin duplicados
        for row in sheet.iter_rows(min_row=2, values_only=True):
            registros_unificados.add(tuple(row))

# Convertir la fecha a formato datetime y luego a string en el formato deseado
for registro in registros_unificados:
    fecha_str = registro[1]  # Supongo que la fecha está en la segunda columna (índice 1)
    fecha_simple = parse_fecha(fecha_str)
    registro_list = list(registro)  # Convertir la tupla a una lista
    registro_list[1] = fecha_simple  # Modificar el elemento deseado
    registro = tuple(registro_list)  # Convertir la lista de nuevo a una tupla

    punion_sheet.append(registro)  # Agregar el registro modificado a la tabla "union"

# Dar formato de tabla a la tabla "union"
tabla_union = openpyxl.worksheet.table.Table(displayName="union", ref=punion_sheet.dimensions)
tabla_union.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                                                     showLastColumn=False, showRowStripes=True,
                                                                     showColumnStripes=True)
punion_sheet.add_table(tabla_union)

# Guardar los cambios en el archivo
workbook.save(file_path)

# Mostrar ventana emergente con mensaje
messagebox.showinfo("Éxito", "La tabla 'union' se ha creado correctamente en la hoja 'punion'.")

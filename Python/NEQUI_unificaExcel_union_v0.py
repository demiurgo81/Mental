import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook

from openpyxl.worksheet.table import Table, TableStyleInfo
def creamodifica_union(workbook, sheet_names):
    def ask_delete_sheet():
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal
        response = messagebox.askyesno("Eliminar hoja", "La hoja 'union' ya existe. ¿Deseas borrarla?")
        root.destroy()  # Destruir la ventana principal después de la decisión
        return response

    def notify_sheet_created():
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal
        messagebox.showinfo("Hoja creada", "La hoja 'union' ha sido creada exitosamente.")
        root.destroy()  # Destruir la ventana principal después de la notificación

    def crear_hoja_union():
        punion_sheet = workbook.create_sheet("union")
        headers = ["Fecha", "Descripción", "Valor", "Saldo"]
        punion_sheet.append(headers)
        print("La tabla 'union' ha sido creada en la hoja 'union'.")
        notify_sheet_created()      
        return punion_sheet

    # Validar la existencia de la hoja "union"
    if "union" in sheet_names:
        if ask_delete_sheet():
            del workbook["union"]
            print("La hoja 'union' ha sido eliminada.")
            punion_sheet = crear_hoja_union()
        else:
            print("Ejecución detenida.")
            return
    else:
        # Crear la hoja "union" y la tabla "union"
        punion_sheet = crear_hoja_union()
    return punion_sheet
def main():
    root = tk.Tk()
    root.withdraw()

    # Solicitar al usuario seleccionar un archivo Excel
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        """
        # Validar la existencia de la hoja "union"
        if "union" in sheet_names:
            response = input("La hoja 'punion' ya existe. ¿Deseas borrarla? (Sí/No): ")
            if response.lower() == "sí":
                workbook.remove(workbook["union"])
                print("La hoja 'punion' ha sido eliminada.")
            else:
                print("Ejecución detenida.")
                return
        else:
            # Crear la hoja "union" y la tabla "union"
            punion_sheet = workbook.create_sheet("union")
            headers = ["Fecha", "Descripción", "Valor", "Saldo"]
            punion_sheet.append(headers)
            print("La tabla 'union' ha sido creada en la hoja 'union'.")
        """
        punion_sheet = creamodifica_union(workbook, sheet_names)

        # Unificar tablas con 9 columnas en la tabla "union"
        union_table = []
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            if sheet.max_column == 4:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    union_table.append(row)

        # Agregar registros a la tabla "union"
        for row in union_table:
            punion_sheet.append(row)

        # Crear la tabla en formato Excel
        tab = Table(displayName="unionT", ref=punion_sheet.dimensions)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        punion_sheet.add_table(tab)

        # Mostrar información sobre la unificación
        print(f"Unificación completada. Total de registros en 'union': {len(union_table)}")

        # Guardar los cambios en el archivo
        workbook.save(file_path)
        print(f"Archivo guardado: {file_path}")

    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()

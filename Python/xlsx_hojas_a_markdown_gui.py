import os
import sys
import math
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import load_workbook
except ImportError:
    # Mensaje claro si ejecutas el .py sin tener openpyxl (para construir el .exe sí debes tenerlo instalado)
    root = tk.Tk(); root.withdraw()
    messagebox.showerror("Dependencia faltante",
                         "No se encontró 'openpyxl'. Instálala con:\n\npip install openpyxl")
    raise

# -------------------- Utilidades --------------------

def get_desktop_dir():
    home = os.path.expanduser("~")
    desk = os.path.join(home, "Desktop")
    return desk if os.path.isdir(desk) else home

def sanitize_filename(name):
    # Quita caracteres no válidos en nombres de archivo
    invalid = '<>:"/\\|?*'
    cleaned = "".join(("_" if c in invalid else c) for c in str(name))
    return cleaned.strip() or "hoja"

def is_empty_cell(v):
    return v is None or (isinstance(v, float) and math.isnan(v)) or (isinstance(v, str) and v.strip() == "")

def escape_md(text):
    if text is None:
        return ""
    s = str(text)
    return s.replace("|", r"\|")

def cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        # ISO legible; ajusta si quieres otro formato
        return v.isoformat()
    return str(v)

def detect_used_range(ws):
    """Devuelve (rows, cols) recortados a las filas/columnas con datos."""
    max_r, max_c = ws.max_row, ws.max_column
    # filas no vacías
    last_row = 0
    for r in range(1, max_r + 1):
        row_has_data = False
        for c in range(1, max_c + 1):
            if not is_empty_cell(ws.cell(r, c).value):
                row_has_data = True
                break
        if row_has_data:
            last_row = r
    if last_row == 0:
        return [], 0  # hoja vacía

    # columnas no vacías
    last_col = 0
    for c in range(1, max_c + 1):
        col_has_data = False
        for r in range(1, last_row + 1):
            if not is_empty_cell(ws.cell(r, c).value):
                col_has_data = True
                break
        if col_has_data:
            last_col = c

    # construir matriz recortada
    rows = []
    for r in range(1, last_row + 1):
        row_vals = []
        for c in range(1, last_col + 1):
            row_vals.append(ws.cell(r, c).value)
        rows.append(row_vals)
    return rows, last_col

def table_markdown(rows, header_first_row=True):
    """rows: lista de filas (listas); header_first_row=True usa la primera fila como encabezado."""
    if not rows:
        return "_(hoja sin datos)_\n"

    # Convierte a str y escapa
    str_rows = [[escape_md(cell_to_str(v)) for v in row] for row in rows]

    if header_first_row and len(str_rows) >= 1:
        header = str_rows[0]
        data = str_rows[1:]
        if not any(h.strip() for h in header):  # si la primera fila está vacía, fuerza encabezados genéricos
            header = [f"col{idx+1}" for idx in range(len(header))]
            data = str_rows
    else:
        # Genera encabezados genéricos
        header = [f"col{idx+1}" for idx in range(len(str_rows[0]))]
        data = str_rows

    header_line = "| " + " | ".join(header) + " |"
    sep_line = "| " + " | ".join(["---"] * len(header)) + " |"
    body_lines = ["| " + " | ".join(row) + " |" for row in data]

    return "\n".join([header_line, sep_line, *body_lines]) + "\n"

# -------------------- Interfaz --------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("XLSX → Markdown (un archivo .md por hoja)")
        self.geometry("640x380")
        self.resizable(False, False)

        self.xlsx_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=get_desktop_dir())
        self.header_first_row = tk.BooleanVar(value=True)
        self.open_folder_after = tk.BooleanVar(value=True)

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        frm_file = ttk.Frame(self); frm_file.pack(fill="x", **pad)
        ttk.Label(frm_file, text="Archivo .xlsx:").pack(side="left")
        ttk.Entry(frm_file, textvariable=self.xlsx_path, width=50).pack(side="left", padx=8)
        ttk.Button(frm_file, text="Seleccionar…", command=self.select_xlsx).pack(side="left")

        frm_out = ttk.Frame(self); frm_out.pack(fill="x", **pad)
        ttk.Label(frm_out, text="Carpeta de salida:").pack(side="left")
        ttk.Entry(frm_out, textvariable=self.output_dir, width=50).pack(side="left", padx=8)
        ttk.Button(frm_out, text="Cambiar…", command=self.select_output_dir).pack(side="left")

        frm_opts = ttk.LabelFrame(self, text="Opciones"); frm_opts.pack(fill="x", **pad)
        ttk.Checkbutton(frm_opts, text="Primera fila es encabezado", variable=self.header_first_row).grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ttk.Checkbutton(frm_opts, text="Abrir carpeta al finalizar", variable=self.open_folder_after).grid(row=0, column=1, sticky="w", padx=8, pady=4)

        frm_actions = ttk.Frame(self); frm_actions.pack(fill="x", **pad)
        ttk.Button(frm_actions, text="Convertir", command=self.convert).pack(side="left")
        ttk.Button(frm_actions, text="Salir", command=self.destroy).pack(side="right")

        frm_log = ttk.LabelFrame(self, text="Progreso"); frm_log.pack(fill="both", expand=True, **pad)
        self.progress = ttk.Progressbar(frm_log, mode="determinate")
        self.progress.pack(fill="x", padx=8, pady=6)
        self.log = tk.Text(frm_log, height=8, state="disabled")
        self.log.pack(fill="both", expand=True, padx=8, pady=6)

    def select_xlsx(self):
        path = filedialog.askopenfilename(
            title="Selecciona un archivo Excel",
            filetypes=[("Excel (*.xlsx)", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        if path:
            self.xlsx_path.set(path)

    def select_output_dir(self):
        d = filedialog.askdirectory(title="Selecciona carpeta de salida", initialdir=get_desktop_dir())
        if d:
            self.output_dir.set(d)

    def log_write(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")
        self.update_idletasks()

    def convert(self):
        xlsx = self.xlsx_path.get().strip()
        out_dir = self.output_dir.get().strip()

        if not xlsx:
            messagebox.showwarning("Falta archivo", "Selecciona un archivo .xlsx.")
            return
        if not os.path.isfile(xlsx):
            messagebox.showerror("Error", "La ruta del .xlsx no existe.")
            return
        if not out_dir:
            messagebox.showwarning("Falta carpeta", "Selecciona una carpeta de salida.")
            return
        if not os.path.isdir(out_dir):
            try:
                os.makedirs(out_dir, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear la carpeta de salida:\n{e}")
                return

        try:
            wb = load_workbook(xlsx, data_only=True, read_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el Excel:\n{e}")
            return

        sheets = wb.sheetnames
        total = len(sheets)
        if total == 0:
            messagebox.showinfo("Sin hojas", "El libro no contiene hojas.")
            return

        self.progress.configure(maximum=total, value=0)
        self.log_write(f"Iniciando conversión: {os.path.basename(xlsx)}")
        base = os.path.splitext(os.path.basename(xlsx))[0]

        ok = 0
        for i, name in enumerate(sheets, start=1):
            self.progress.configure(value=i-1)
            self.update_idletasks()

            ws = wb[name]
            rows, _ = detect_used_range(ws)

            md = f"# {name}\n\n" + table_markdown(rows, header_first_row=self.header_first_row.get())
            sheet_fname = f"{sanitize_filename(base)}__{sanitize_filename(name)}.md"
            out_path = os.path.join(out_dir, sheet_fname)

            try:
                with open(out_path, "w", encoding="utf-8", newline="\n") as f:
                    f.write(md)
                ok += 1
                self.log_write(f"✔ Guardado: {out_path}")
            except Exception as e:
                self.log_write(f"✖ Error guardando {sheet_fname}: {e}")

            self.progress.configure(value=i)
            self.update_idletasks()

        self.log_write(f"Finalizado. {ok}/{total} archivos .md generados.")
        messagebox.showinfo("Completado", f"Generados {ok} de {total} archivos .md en:\n{out_dir}")

        if self.open_folder_after.get():
            try:
                if sys.platform.startswith("win"):
                    os.startfile(out_dir)  # type: ignore[attr-defined]
                elif sys.platform == "darwin":
                    os.system(f'open "{out_dir}"')
                else:
                    os.system(f'xdg-open "{out_dir}"')
            except Exception:
                pass


if __name__ == "__main__":
    app = App()
    app.mainloop()
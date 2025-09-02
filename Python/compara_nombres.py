import openpyxl
from openpyxl import Workbook
import re
import os
from difflib import SequenceMatcher
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox

def clean_and_prepare_string(s):
    if s is None:
        return ""
    s = s.upper()
    # Reemplazar letras acentuadas por sus equivalentes sin acento
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U', 'Ñ': 'N'
    }
    for accented_char, unaccented_char in replacements.items():
        s = s.replace(accented_char, unaccented_char)
    # Eliminar caracteres no deseados
    s = re.sub(r'[^A-Z0-9 ÁÉÍÓÚÜÑ]', '', s)
    # Reducir espacios consecutivos a uno solo
    s = re.sub(r'\s+', ' ', s)
    # Eliminar espacios al inicio y al final
    s = s.strip()
    return s


def compare_strings(s1, s2):
    s1_words = s1.split(' ')
    s2_words = s2.split(' ')
    s1_set = set(s1_words)
    s2_set = set(s2_words)
    common_words = s1_set.intersection(s2_set)
    total_words = s1_set.union(s2_set)
    similarity_ratio = len(common_words) / len(total_words)
    return similarity_ratio


def find_column_index(sheet, column_name):
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == column_name:
            return col[0].column
    raise ValueError(f"Column {column_name} not found in sheet")

def find_most_similar(workbook, hoja1, hoja2, campo1, campo2):
    ws1 = workbook[hoja1]
    ws2 = workbook[hoja2]

    col1_index = find_column_index(ws1, campo1)
    col2_index = find_column_index(ws2, campo2)
    
    col1 = [clean_and_prepare_string(ws1.cell(row=row, column=col1_index).value) for row in range(2, ws1.max_row + 1)]
    col2 = [clean_and_prepare_string(ws2.cell(row=row, column=col2_index).value) for row in range(2, ws2.max_row + 1)]
    
    results = []
    
    root = tk.Tk()
    root.title("Comparación de registros")
    
    progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=20)
    
    label = tk.Label(root, text="Procesando comparaciones...")
    label.pack(pady=10)
    
    max_value = len(col1) * len(col2)
    progress['maximum'] = max_value
    
    def update_progress(current):
        progress['value'] = current
        root.update()
    
    current_progress = 0
    
    for item1 in col1:
        best_match = None
        second_best_match = None
        best_similarity = 0
        second_best_similarity = 0
        for item2 in col2:
            similarity = compare_strings(item1, item2)
            if similarity > best_similarity:
                second_best_match = best_match
                second_best_similarity = best_similarity
                best_similarity = similarity
                best_match = item2
            elif similarity > second_best_similarity:
                second_best_similarity = similarity
                second_best_match = item2
            current_progress += 1
            update_progress(current_progress)
        results.append((item1, best_match, best_similarity, second_best_match, second_best_similarity))
    
    root.destroy()
    
    #result_ws = workbook.create_sheet(title="resultado")

    #nuevo excel con solo el resultado
    workbook = Workbook()
    result_ws = workbook['Sheet']
    result_ws.append([campo1, "Mejor Coincidencia", "Índice de Similitud", "Segunda Mejor Coincidencia", "Segundo Índice de Similitud"])
    
    for res in results:
        result_ws.append(res)
    
    return workbook

def centrar_ventana(root):
    root.update_idletasks()
    ancho = root.winfo_width()
    alto = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (ancho // 2)
    y = (root.winfo_screenheight() // 2) - (alto // 2)
    root.geometry(f'{ancho}x{alto}+{x}+{y}')

def seleccionar_hojas(workbook):
    def confirmar_seleccion():
        hoja1 = combo_hoja1.get()
        hoja2 = combo_hoja2.get()
        if hoja1 in hojas and hoja2 in hojas:
            seleccionadas.append(hoja1)
            seleccionadas.append(hoja2)
            root.destroy()
        else:
            messagebox.showerror("Error", "Las hojas seleccionadas no son válidas.")

    root = tk.Tk()
    root.title("Seleccionar Hojas")
    root.attributes('-topmost', True)  # Llevar la ventana al frente

    hojas = workbook.sheetnames
    seleccionadas = []

    tk.Label(root, text="Seleccione la primera hoja:").pack(pady=5)
    combo_hoja1 = ttk.Combobox(root, values=hojas)
    combo_hoja1.pack(pady=5)

    tk.Label(root, text="Seleccione la segunda hoja:").pack(pady=5)
    combo_hoja2 = ttk.Combobox(root, values=hojas)
    combo_hoja2.pack(pady=5)

    tk.Button(root, text="Confirmar", command=confirmar_seleccion).pack(pady=20)

    centrar_ventana(root)
    root.mainloop()

    if len(seleccionadas) == 2:
        return seleccionadas[0], seleccionadas[1]
    else:
        return None, None

def seleccionar_campos(workbook, hoja1, hoja2):
    def confirmar_seleccion():
        campo1 = combo_campo1.get()
        campo2 = combo_campo2.get()
        if campo1 in campos_hoja1 and campo2 in campos_hoja2:
            seleccionados.append(campo1)
            seleccionados.append(campo2)
            root.destroy()
        else:
            messagebox.showerror("Error", "Los campos seleccionados no son válidos.")

    root = tk.Tk()
    root.title("Seleccionar Campos")
    root.attributes('-topmost', True)  # Llevar la ventana al frente

    campos_hoja1 = [cell.value for cell in workbook[hoja1][1]]
    campos_hoja2 = [cell.value for cell in workbook[hoja2][1]]
    seleccionados = []

    tk.Label(root, text=f"Seleccione un campo de la hoja {hoja1}:").pack(pady=5)
    combo_campo1 = ttk.Combobox(root, values=campos_hoja1)
    combo_campo1.pack(pady=5)

    tk.Label(root, text=f"Seleccione un campo de la hoja {hoja2}:").pack(pady=5)
    combo_campo2 = ttk.Combobox(root, values=campos_hoja2)
    combo_campo2.pack(pady=5)

    tk.Button(root, text="Confirmar", command=confirmar_seleccion).pack(pady=20)

    centrar_ventana(root)
    root.mainloop()

    if len(seleccionados) == 2:
        return seleccionados[0], seleccionados[1]
    else:
        return None, None

def try_open_workbook(file_path):
    while True:
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except PermissionError:
            root = tk.Tk()
            root.withdraw()  # Ocultar la ventana principal
            messagebox.showwarning("Archivo en uso", f"El archivo '{file_path}' está en uso. Por favor, ciérralo y presiona 'Aceptar' para continuar.")
            root.destroy()  # Destruir la ventana principal después de cerrar el mensaje

# Solicitar al usuario seleccionar un archivo Excel
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

workbook = try_open_workbook(file_path)
sheet_names = workbook.sheetnames
print(sheet_names)

hoja1, hoja2 = seleccionar_hojas(workbook)
print("selecciono la hoja: "+hoja1+" y  "+hoja2+" del archivo excel")

if hoja1 and hoja2:
    campo1, campo2 = seleccionar_campos(workbook, hoja1, hoja2)
    print("eligio el campo: "+campo1+" de la hoja: "+hoja1+" y el campo: "+campo2+" de la hoja: "+hoja2+" del archivo excel")
    if campo1 and campo2:
        workbook = find_most_similar(workbook, hoja1, hoja2, campo1, campo2)
        #workbook.save(file_path)
        #genera el resultado en un archivo diferente.
        workbook.save(os.path.dirname(file_path)+"/ResultadoComparacion.xlsx")
        os.startfile(os.path.dirname(file_path)+"/ResultadoComparacion.xlsx")
        print(f"Hoja 1: {hoja1}, Campo 1: {campo1}")
        print(f"Hoja 2: {hoja2}, Campo 2: {campo2}")
    else:
        print("No se encontraron los campos")
else:
    print("NO se encontraron las hojas")

import cx_Oracle
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from fuzzywuzzy import process
import warnings



def main():
    root = tk.Tk()
    root.withdraw()
    try:
        # Eliminar la advertencia
        warnings.filterwarnings("ignore", category=UserWarning, module='fuzzywuzzy')
        # Datos de conexión
        dsn_tns = cx_Oracle.makedsn("192.168.18.198", "1523", service_name="srvctl")
        connection = cx_Oracle.connect(user="USR_EAP_LOGS", password="3oKF$72Cvtf$", dsn=dsn_tns)

        # Ejecutar la consulta
        cursor = connection.cursor()
        cursor.execute("SELECT    id,    nombre FROM    df_proveedor")

        # Obtener los nombres de las columnas
        columns = [col[0] for col in cursor.description]

        # Obtener los datos
        data = cursor.fetchall()

        # Cerrar la conexión
        cursor.close()
        connection.close()
        

        def actualizar_excel_con_ids(workbook,columns, data):
                                      
            sheet = workbook['proveedores']

            # Obtener los valores de la columna 'Empresa' en la hoja 'proveedores'
            empresas = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]

            # Crear un diccionario con los nombres e ids obtenidos de la consulta a la base de datos
            nombre_id_dict = {nombre.upper(): id_ for id_, nombre in data}

            # Iterar sobre las empresas y encontrar la mejor coincidencia
            for idx, empresa in enumerate(empresas, start=2):
                if empresa:
                    empresa_upper = empresa.upper()
                    best_match, score = process.extractOne(empresa_upper, nombre_id_dict.keys())

                    if best_match and score > 70:  # Umbral de similitud más bajo
                        best_id = nombre_id_dict[best_match]
                        sheet.cell(row=idx, column=2).value = best_id  
                    else:
                        # Verificar coincidencia de una sola palabra
                        empresa_words = set(empresa_upper.split())
                        for nombre in nombre_id_dict.keys():
                            nombre_words = set(nombre.split())
                            if empresa_words & nombre_words:  # Si hay alguna palabra en común
                                sheet.cell(row=idx, column=3).value = nombre_id_dict[nombre]
                                break
                        else:
                            sheet.cell(row=idx, column=3).value = 'NOT FOUND'
                else:
                    sheet.cell(row=idx, column=3).value = 'NOT FOUND'

            # Guardar el archivo Excel actualizado
            workbook.save(file_path)

        # Solicitar al usuario seleccionar un archivo Excel
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        print(sheet_names)
 
        # Llamar a la función
        actualizar_excel_con_ids(workbook,columns, data)
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()
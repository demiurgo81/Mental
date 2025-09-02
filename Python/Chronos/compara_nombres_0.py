import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from fuzzywuzzy import process
from difflib import SequenceMatcher

import warnings

def obtener_columna_por_nombre(hoja, nombre_columna):
    for col in hoja.iter_cols(1, hoja.max_column):
        if col[0].value == nombre_columna:
            return col
    return None

def comparar_y_guardar_resultados(workbook, hoja1, hoja2, campo1, campo2, campo_id):
    hoja_1 = workbook[hoja1]
    hoja_2 = workbook[hoja2]

    columna_hoja1 = obtener_columna_por_nombre(hoja_1, campo1)
    columna_hoja2 = obtener_columna_por_nombre(hoja_2, campo2)
    columna_id_hoja2 = obtener_columna_por_nombre(hoja_2, campo_id)

    if columna_hoja1 is None or columna_hoja2 is None or columna_id_hoja2 is None:
        raise ValueError("Una o más columnas no se encontraron.")

    #resultados = [["Campo 1 Original", "Campo 1", "Campo 2 Original", "Campo 2", "ID", "Estado"]]
    resultados = [["Campo 1 Original", "Campo 2 Original", "ID", "Estado"]]

    for celda1 in columna_hoja1[1:]:
        valor1_original = celda1.value
        valor1 = str(valor1_original).upper()
        mejor_similitud = 0
        mejor_valor2 = ""
        mejor_valor2_original = ""
        mejor_id = ""
        encontrado = False

        for celda2, celda_id in zip(columna_hoja2[1:], columna_id_hoja2[1:]):
            valor2_original = celda2.value
            valor2 = str(valor2_original).upper()
            similitud = SequenceMatcher(None, valor1, valor2).ratio()

            if similitud > mejor_similitud and similitud > 0.7:
                mejor_similitud = similitud
                mejor_valor2 = valor2
                mejor_valor2_original = valor2_original
                mejor_id = celda_id.value
                resultado = similitud
                encontrado = True

        if not encontrado:
            for celda2, celda_id in zip(columna_hoja2[1:], columna_id_hoja2[1:]):
                valor2_original = celda2.value
                valor2 = str(valor2_original).upper()
                if any(palabra in valor2.split() for palabra in valor1.split()):
                    mejor_valor2 = valor2
                    mejor_valor2_original = valor2_original
                    mejor_id = celda_id.value
                    resultado = 0.5
                    encontrado = True
                    break

        if encontrado:
            resultados.append([valor1_original, mejor_valor2_original, mejor_id, resultado])#envia similitud o pioresnada EN LUGAR DE MATCH
        else:
            resultados.append([valor1_original, "", "", 0])

    if "resultado" in workbook.sheetnames:
        resultado_hoja = workbook["resultado"]
        workbook.remove(resultado_hoja)

    resultado_hoja = workbook.create_sheet("resultado")

    for fila in resultados:
        resultado_hoja.append(fila)

    workbook.save(file_path)

def centrar_ventana(root):
    root.update_idletasks()
    ancho = root.winfo_width()
    alto = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (ancho // 2)
    y = (root.winfo_screenheight() // 2) - (alto // 2)
    root.geometry(f'{ancho}x{alto}+{x}+{y}')

def seleccionar_hojas(workbook):
    def confirmar_seleccion():
        hoja1 = combo_hoja1.get()
        hoja2 = combo_hoja2.get()
        if hoja1 in hojas and hoja2 in hojas:
            seleccionadas.append(hoja1)
            seleccionadas.append(hoja2)
            root.destroy()
        else:
            messagebox.showerror("Error", "Las hojas seleccionadas no son válidas.")

    root = tk.Tk()
    root.title("Seleccionar Hojas")
    root.attributes('-topmost', True)  # Llevar la ventana al frente

    hojas = workbook.sheetnames
    seleccionadas = []

    tk.Label(root, text="Seleccione la primera hoja:").pack(pady=5)
    combo_hoja1 = ttk.Combobox(root, values=hojas)
    combo_hoja1.pack(pady=5)

    tk.Label(root, text="Seleccione la segunda hoja:").pack(pady=5)
    combo_hoja2 = ttk.Combobox(root, values=hojas)
    combo_hoja2.pack(pady=5)

    tk.Button(root, text="Confirmar", command=confirmar_seleccion).pack(pady=20)

    centrar_ventana(root)
    root.mainloop()

    if len(seleccionadas) == 2:
        return seleccionadas[0], seleccionadas[1]
    else:
        return None, None

def seleccionar_campos(workbook, hoja1, hoja2):
    def confirmar_seleccion():
        campo1 = combo_campo1.get()
        campo2 = combo_campo2.get()
        if campo1 in campos_hoja1 and campo2 in campos_hoja2:
            seleccionados.append(campo1)
            seleccionados.append(campo2)
            root.destroy()
        else:
            messagebox.showerror("Error", "Los campos seleccionados no son válidos.")

    root = tk.Tk()
    root.title("Seleccionar Campos")
    root.attributes('-topmost', True)  # Llevar la ventana al frente

    campos_hoja1 = [cell.value for cell in workbook[hoja1][1]]
    campos_hoja2 = [cell.value for cell in workbook[hoja2][1]]
    seleccionados = []

    tk.Label(root, text=f"Seleccione un campo de la hoja {hoja1}:").pack(pady=5)
    combo_campo1 = ttk.Combobox(root, values=campos_hoja1)
    combo_campo1.pack(pady=5)

    tk.Label(root, text=f"Seleccione un campo de la hoja {hoja2}:").pack(pady=5)
    combo_campo2 = ttk.Combobox(root, values=campos_hoja2)
    combo_campo2.pack(pady=5)

    tk.Button(root, text="Confirmar", command=confirmar_seleccion).pack(pady=20)

    centrar_ventana(root)
    root.mainloop()

    if len(seleccionados) == 2:
        return seleccionados[0], seleccionados[1]
    else:
        return None, None

def try_open_workbook(file_path):
    while True:
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except PermissionError:
            root = tk.Tk()
            root.withdraw()  # Ocultar la ventana principal
            messagebox.showwarning("Archivo en uso", f"El archivo '{file_path}' está en uso. Por favor, ciérralo y presiona 'Aceptar' para continuar.")
            root.destroy()  # Destruir la ventana principal después de cerrar el mensaje

# Solicitar al usuario seleccionar un archivo Excel
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

workbook = try_open_workbook(file_path)
sheet_names = workbook.sheetnames
print(sheet_names)

hoja1, hoja2 = seleccionar_hojas(workbook)
print("selecciono la hoja: "+hoja1+" y  "+hoja2+" del archivo excel")

if hoja1 and hoja2:
    campo1, campo2 = seleccionar_campos(workbook, hoja1, hoja2)
    print("eligio el campo: "+campo1+" de la hoja: "+hoja1+" y el campo: "+campo2+" de la hoja: "+hoja2+" del archivo excel")
    
    if campo1 and campo2:
        comparar_y_guardar_resultados(workbook, hoja1, hoja2, campo1, campo2,"ID")
        print(f"Hoja 1: {hoja1}, Campo 1: {campo1}")
        print(f"Hoja 2: {hoja2}, Campo 2: {campo2}")
    else:
        print("no se encontraron los campos")
else:
    print("NO se encontraron las hojas")

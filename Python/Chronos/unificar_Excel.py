import openpyxl
from tkinter import filedialog, Tk

# Solicitar al usuario que seleccione un archivo Excel
root = Tk()
root.withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

# Cargar el archivo Excel
workbook = openpyxl.load_workbook(file_path)

# Crear una nueva hoja llamada "punion"
punion_sheet = workbook.create_sheet("unificado")

# Buscar todas las tablas en el archivo
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    # Copiar los registros a la tabla "unificado"
    for row in sheet.iter_rows(min_row=1, values_only=True):
        punion_sheet.append(row)

# Guardar los cambios en el archivo
workbook.save(file_path)
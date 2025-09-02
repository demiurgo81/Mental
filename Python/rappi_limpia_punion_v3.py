import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def abrir_archivo():
    archivo = filedialog.askopenfilename(
        initialdir="/",
        title="Selecciona un archivo Excel",
        filetypes=(("Ficheros Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
    )
    return archivo

def validar_hoja_punion(archivo):
    try:
        df = pd.read_excel(archivo, sheet_name="punion")
        columnas_requeridas = ["tipo", "fecha", "descripcion", "valor", "capital", "cuotas", "pendiente", "tasaMV", "tasaEA"]
        if all(col in df.columns for col in columnas_requeridas):
            return df
        else:
            return None
    except Exception as e:
        return None

def procesar_descripcion(df):
    for i, row in df.iterrows():
        descripcion_actual = row["descripcion"]
        if pd.isna(descripcion_actual):
            if i > 0:
                fecha_anterior = df.loc[i - 1, "fecha"]
                valor_anterior = df.loc[i - 1, "valor"]
                if pd.isna(fecha_anterior) and pd.isna(valor_anterior):
                    descripcion_anterior = df.loc[i - 1, "descripcion"]
                    descripcion_siguiente = df.loc[i + 1, "descripcion"]
                    nueva_descripcion = f"{descripcion_anterior} {descripcion_siguiente}"
                    df.at[i, "descripcion"] = nueva_descripcion
    return df

def eliminar_registros_vacios(df):
    # Eliminar registros con campo "descripcion" diferente de vacío
    df.dropna(subset=["descripcion"], inplace=True)

    # Eliminar registros con más de 5 campos vacíos
    df.dropna(thresh=len(df.columns) - 5, inplace=True)

    return df

def clean_fecha(df):
    # Expresión regular para la estructura '([1-2])([0-9]{3})(-)([0-1])([0-9])(-)([0-3])([0-9])'
    regex = r"([1-2])([0-9]{3})(-)([0-1])([0-9])(-)([0-3])([0-9])"

    # Convertir el campo "fecha" al formato de fecha con formato explícito
    df["fecha"] = pd.to_datetime(df["fecha"], format="%Y-%m-%d", errors="coerce")

    # Formatear la columna "fecha" como cadena con el formato deseado
    df["fecha"] = df["fecha"].dt.strftime("%Y-%m-%d")

    # Eliminar registros con valores de fecha vacíos
    df_cleaned = df.dropna(subset=["fecha"])

    return df_cleaned

def rectifica_valores(df):
    
    # Reemplazar todos los valores en "cuotas" por 0
    df.loc[:, "cuotas"] = 0
    
    # Función para convertir el formato monetario
    def convert_currency(value):
        value = str(value)
        # Si el valor tiene comas como separador de miles y puntos como separador decimal
        if '.' in value and ',' in value:
            value = value.replace('.', '').replace(',', '.')
        else:
            value = value.replace(',', '')
        value = re.sub(r'[^\d.-]', '', value)
        try:
            return float(value) if value else 0
        except ValueError:
            return 0
    
    # Función para convertir el formato porcentual
    def convert_percentage(value):
        value = str(value)
        value = value.replace('%', '').replace(',', '.')
        try:
            return float(value) if value else 0
        except ValueError:
            return 0
    
    # Rectificar formato de valores en "valor", "capital" y "pendiente"
    for col in ["valor", "capital", "pendiente"]:
        df.loc[:, col] = df[col].apply(convert_currency)
    
    # Rectificar formato de valores en "tasaMV" y "tasaEA"
    for col in ["tasaMV", "tasaEA"]:
        df.loc[:, col] = df[col].apply(convert_percentage)
    
    return df

def mostrar_dataframe(df):
    # Crear ventana emergente
    ventana = tk.Tk()
    ventana.title("Tabla del DataFrame")
    
    # Crear tabla con el contenido del DataFrame
    tabla = tk.Text(ventana, wrap=tk.NONE)
    tabla.insert("1.0", df.to_string(index=False))
    tabla.pack()
    
    # Crear botón "Continuar"
    def continuar():
        ventana.destroy()  # Cerrar la ventana emergente
        # Aquí puedes reanudar la ejecución del script
    
    boton_continuar = tk.Button(ventana, text="Continuar", command=continuar)
    boton_continuar.pack()
    
    # Mostrar la ventana emergente al frente de todos los programas
    ventana.attributes("-topmost", True)
    ventana.mainloop()

def main():
    archivo_seleccionado = abrir_archivo()
    if not archivo_seleccionado:
        print("No se seleccionó ningún archivo.")
        return

    df_punion = validar_hoja_punion(archivo_seleccionado)
    if df_punion is None:
        print("La hoja 'punion' no existe o no contiene los encabezados requeridos.")
        return

    df_punion_procesado = procesar_descripcion(df_punion)
    df_punion_sinvacios = eliminar_registros_vacios(df_punion_procesado)
    df_punion_fixingfecha = clean_fecha(df_punion_sinvacios)
    df_punion_formatovalores = rectifica_valores(df_punion_fixingfecha)
    df_punion_final = df_punion_formatovalores

#    mostrar_dataframe(df_punion_formatovalores)

    # Carga el archivo Excel existente
    workbook = openpyxl.load_workbook(archivo_seleccionado)

    # Obtén la hoja de cálculo existente o crea una nueva si no existe
    
    if "punion" in workbook.sheetnames  :
        del workbook['punion']
        workbook.create_sheet("punion")
    else: 
        workbook.create_sheet("punion")
    hoja_union = workbook["punion"]
    # Escribe los datos del DataFrame en la hoja de cálculo
    for row in dataframe_to_rows(df_punion_final, index=False, header=True):
        hoja_union.append(row)

    # Guarda los cambios en el archivo Excel
    workbook.save(archivo_seleccionado)
    

    print("Cambios realizados y guardados en el archivo Excel.")

if __name__ == "__main__":
    main()

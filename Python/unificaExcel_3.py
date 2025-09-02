import tkinter as tk
from tkinter import filedialog
import openpyxl

def main():
    root = tk.Tk()
    root.withdraw()

    # Solicitar al usuario seleccionar un archivo Excel
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames

        # Mostrar información sobre las tablas encontradas
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            num_rows = sheet.max_row
            num_columns = sheet.max_column
            print(f"Tabla '{sheet_name}': {num_rows} registros, {num_columns} columnas")

        # Validar la existencia de la hoja "punion"
        if "punion" in sheet_names:
            response = input("La hoja 'punion' ya existe. ¿Deseas borrarla? (Sí/No): ")
            if response.lower() == "sí":
                workbook.remove(workbook["punion"])
                print("La hoja 'punion' ha sido eliminada.")
            else:
                print("Ejecución detenida.")
        else:
            # Crear la hoja "punion" y la tabla "union"
            punion_sheet = workbook.create_sheet("punion")
            headers = ["tipo", "fecha", "descripcion", "valor", "capital", "cuotas", "pendiente", "tasaMV", "tasaEA"]
            punion_sheet.append(headers)
            print("La tabla 'union' ha sido creada en la hoja 'punion'.")

            # Unificar tablas con 9 columnas en la tabla "union"
            union_table = []
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                if sheet.max_column == 9:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        union_table.append(row)

            # Agregar registros a la tabla "union"
            for row in union_table:
                punion_sheet.append(row)

            # Mostrar información sobre la unificación
            print(f"Unificación completada. Total de registros en 'union': {len(union_table)}")

        # Guardar los cambios en el archivo
        workbook.save(file_path)
        print(f"Archivo guardado: {file_path}")

    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def abrir_archivo():
    archivo = filedialog.askopenfilename(
        initialdir="/",
        title="Selecciona un archivo Excel",
        filetypes=(("Ficheros Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
    )
    return archivo

def validar_hoja_punion(archivo):
    try:
        df = pd.read_excel(archivo, sheet_name="punion")
        columnas_requeridas = ["tipo", "fecha", "descripcion", "valor", "capital", "cuotas", "pendiente", "tasaMV", "tasaEA"]
        if all(col in df.columns for col in columnas_requeridas):
            return df
        else:
            return None
    except Exception as e:
        return None

def procesar_descripcion(df):
    for i, row in df.iterrows():
        descripcion_actual = row["descripcion"]
        if pd.isna(descripcion_actual):
            if i > 0:
                fecha_anterior = df.loc[i - 1, "fecha"]
                valor_anterior = df.loc[i - 1, "valor"]
                if pd.isna(fecha_anterior) and pd.isna(valor_anterior):
                    descripcion_anterior = df.loc[i - 1, "descripcion"]
                    descripcion_siguiente = df.loc[i + 1, "descripcion"]
                    nueva_descripcion = f"{descripcion_anterior} {descripcion_siguiente}"
                    df.at[i, "descripcion"] = nueva_descripcion
    return df

def eliminar_registros_vacios(df):
    # Eliminar registros con campo "descripcion" diferente de vacío
    df.dropna(subset=["descripcion"], inplace=True)

    # Eliminar registros con más de 5 campos vacíos
    df.dropna(thresh=len(df.columns) - 5, inplace=True)

    return df

def clean_fecha(df):
    # Expresión regular para la estructura '([1-2])([0-9]{3})(-)([0-1])([0-9])(-)([0-3])([0-9])'
    regex = r"([1-2])([0-9]{3})(-)([0-1])([0-9])(-)([0-3])([0-9])"

    # Convertir el campo "fecha" al formato de fecha
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")

    # Formatear la columna "fecha" como cadena con el formato deseado
    df["fecha"] = df["fecha"].dt.strftime("%Y-%m-%d")

    return df

def procesar_dataframe(df):
    # Elimina los signos de moneda y convierte a números
    df['valor'] = df['valor'].str.replace('[\$,€]', '', regex=True).astype(float)
    df['capital'] = df['capital'].str.replace('[\$,€]', '', regex=True).astype(float)
    df['pendiente'] = df['pendiente'].str.replace('[\$,€]', '', regex=True).astype(float)

    # Formatea los números con coma como separador decimal
    df['valor'] = df['valor'].apply(lambda x: f'{x:,.2f}')
    df['capital'] = df['capital'].apply(lambda x: f'{x:,.2f}')
    df['pendiente'] = df['pendiente'].apply(lambda x: f'{x:,.2f}')

    return df

def mostrar_dataframe(df):
    # Crear ventana emergente
    ventana = tk.Tk()
    ventana.title("Tabla del DataFrame")
    
    # Crear tabla con el contenido del DataFrame
    tabla = tk.Text(ventana, wrap=tk.NONE)
    tabla.insert("1.0", df.to_string(index=False))
    tabla.pack()
    
    # Crear botón "Continuar"
    def continuar():
        ventana.destroy()  # Cerrar la ventana emergente
        # Aquí puedes reanudar la ejecución del script
    
    boton_continuar = tk.Button(ventana, text="Continuar", command=continuar)
    boton_continuar.pack()
    
    # Mostrar la ventana emergente al frente de todos los programas
    ventana.attributes("-topmost", True)
    ventana.mainloop()

def main():
    archivo_seleccionado = abrir_archivo()
    if not archivo_seleccionado:
        print("No se seleccionó ningún archivo.")
        return

    df_punion = validar_hoja_punion(archivo_seleccionado)
    if df_punion is None:
        print("La hoja 'punion' no existe o no contiene los encabezados requeridos.")
        return

    df_punion_procesado = procesar_descripcion(df_punion)
    df_punion_sinvacios = eliminar_registros_vacios(df_punion_procesado)
    df_punion_fixingfecha = clean_fecha(df_punion_sinvacios)
    
    df_punion_final = df_punion_fixingfecha
    mostrar_dataframe(df_punion_final)

#    df_punion_final = procesar_dataframe(df_punion_valores)

    # Carga el archivo Excel existente
#    workbook = openpyxl.load_workbook(archivo_seleccionado)

    # Obtén la hoja de cálculo existente o crea una nueva si no existe
    
#    if "punion" in workbook.sheetnames  :
#        del workbook['punion']
#        workbook.create_sheet("punion")
#    else: 
#        workbook.create_sheet("punion")
#    hoja_union = workbook["punion"]
    # Escribe los datos del DataFrame en la hoja de cálculo
#    for row in dataframe_to_rows(df_punion_final, index=False, header=True):
#        hoja_union.append(row)
#
    # Guarda los cambios en el archivo Excel
#    workbook.save(archivo_seleccionado)

    print("Cambios realizados y guardados en el archivo Excel.")

if __name__ == "__main__":
    main()
